# -*- coding: utf-8 -*-
"""
Объединение Excel-файлов и постобработка.

Имена входных файлов и дата "поточнадата" берутся из app_settings.json.
Логические блоки (статус2БЛОК, S070БЛОК, S186/190/242БЛОК) — из status2_map.json.

Листы:
- Лист1 — полная таблица + КомКредСумаУзвітномуперіоді; прогресс каждые 40 сек
- Выборка — ТОЛЬКО строки, где col_14 > 0; сумма=sum(col_14), колличество=count
- ДляНБУ — как «Выборка», но доп. фильтр сумма>0 и колличество>0
- КомисссияПоКредитамВсе — агрегат (без S186* и S242*), метрики:
    КомКредСумаУзвітномуперіоді = sum(col_29 - col_13 + col_22 - col_28)
    СумаНазвітнудату            = sum(col_34)
- КомисссияПоКредитамНБУ — агрегат из предыдущего листа; удаляем строки, где обе суммы == 0
"""

import os
import re
import json
import time
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook

# === Пути к файлам настроек ===
SETTINGS_JSON = "app_settings.json"   # тут файлы и поточнадата (+ sheet_name опционально)
CONFIG_JSON   = "status2_map.json"    # тут блоки правил/отображений

# === Константы по Excel ===
SHEET_NAME = "Лист1"  # может быть переопределён в app_settings.json -> "sheet_name"

# === Форматы колонок ===
TEXT_COLS = [1, 2, 3, 4, 5, 6, 37]                        # текст
DATE_COLS = [7, 8, 36]                                    # dd.mm.yyyy
INT_OR_EMPTY_COLS = [9]                                   # целое или пусто
NUMERIC_ZERO_IF_EMPTY_COLS = list(range(10, 36)) + [38]   # число или 0
ALL_COLS_1BASED = list(range(1, 39))
COL_NAMES = [f"col_{i}" for i in ALL_COLS_1BASED]

# === Утилиты нормализации ===
def is_empty_like(val) -> bool:
    if pd.isna(val):
        return True
    if isinstance(val, str):
        s = val.strip().lower()
        return s == "" or s == "-" or s == "null"
    return False

def norm_text(val) -> str:
    if is_empty_like(val):
        return ""
    if isinstance(val, (int, np.integer)):
        return str(val)
    if isinstance(val, (float, np.floating)):
        if not np.isfinite(val):
            return ""
        if float(val).is_integer():
            return str(int(val))
        return format(val, "f").rstrip("0").rstrip(".")
    return str(val)

def norm_date(val) -> str:
    if is_empty_like(val):
        return ""
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.strftime("%d.%m.%Y")
    try:
        if isinstance(val, (int, float)) and not np.isnan(val):
            ts = pd.to_datetime("1899-12-30") + pd.to_timedelta(int(val), unit="D")
            return ts.strftime("%d.%m.%Y")
    except Exception:
        pass
    try:
        s = str(val).strip().replace("/", ".").replace("-", ".")
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        return "" if pd.isna(dt) else dt.strftime("%d.%m.%Y")
    except Exception:
        return ""

def to_int_or_empty(val):
    if is_empty_like(val):
        return ""
    try:
        s = str(val).strip().replace(" ", "").replace(",", ".")
        num = float(s)
        if np.isnan(num):
            return ""
        return int(round(num))
    except Exception:
        return ""

def to_number_zero_if_empty(val):
    if is_empty_like(val):
        return 0
    try:
        s = str(val).strip().replace(" ", "").replace(",", ".")
        num = float(s)
        return 0 if np.isnan(num) else num
    except Exception:
        return 0

def safe_num(x):
    try:
        return float(x)
    except Exception:
        return 0.0

# === Чтение заголовков исходника (первая строка) ===
def get_headers(path: str, sheet_name: str):
    try:
        headers_df = pd.read_excel(path, sheet_name=sheet_name, usecols=list(range(0, 38)),
                                   header=None, nrows=1, engine="openpyxl")
    except ValueError:
        headers_df = pd.read_excel(path, sheet_name=0, usecols=list(range(0, 38)),
                                   header=None, nrows=1, engine="openpyxl")
    headers = [norm_text(v) for v in headers_df.iloc[0].tolist()]
    if len(headers) < 38: headers += [""] * (38 - len(headers))
    elif len(headers) > 38: headers = headers[:38]
    return headers

# === Чтение и приведение одного файла ===
def read_and_process_one(path: str, sheet_name: str) -> pd.DataFrame:
    skip_rows = 2 if "MC_NBU" in os.path.basename(path) else 1
    try:
        df_raw = pd.read_excel(path, sheet_name=sheet_name, usecols=list(range(0, 38)),
                               header=None, engine="openpyxl", skiprows=skip_rows)
    except ValueError:
        df_raw = pd.read_excel(path, sheet_name=0, usecols=list(range(0, 38)),
                               header=None, engine="openpyxl", skiprows=skip_rows)
    df_raw.columns = COL_NAMES[: df_raw.shape[1]]

    # Обрезка по первой пустой в кол.1
    col1_empty = df_raw["col_1"].apply(is_empty_like)
    first_empty_idx = col1_empty.idxmax() if col1_empty.any() else None
    df = df_raw.iloc[:first_empty_idx, :].copy() if (first_empty_idx is not None and col1_empty[first_empty_idx]) else df_raw.copy()

    # Выровнять колонки
    for c in COL_NAMES:
        if c not in df.columns:
            df[c] = ""
    df = df[COL_NAMES]

    # Приведение типов
    for i in TEXT_COLS:
        df[f"col_{i}"] = df[f"col_{i}"].map(norm_text)
    for i in DATE_COLS:
        df[f"col_{i}"] = df[f"col_{i}"].map(norm_date)
    for i in INT_OR_EMPTY_COLS:
        df[f"col_{i}"] = df[f"col_{i}"].map(to_int_or_empty)
    for i in NUMERIC_ZERO_IF_EMPTY_COLS:
        df[f"col_{i}"] = df[f"col_{i}"].map(to_number_zero_if_empty)
    return df

# === Вспомогательное ===
def extract_date_str(filename: str) -> str:
    m = re.search(r"_(\d{4}-\d{2})", filename)
    return m.group(1) if m else "unknown"

def parse_status2_entry(entry):
    if isinstance(entry, str):
        return entry, False
    if isinstance(entry, dict):
        val = entry.get("value", entry.get("значение", ""))
        flag = str(entry.get("проверкаколонка38", "")).strip().lower() == "да"
        return val, flag
    return "", False

def parse_s070_block(block_dict):
    code_map, label_map = {}, {}
    if not isinstance(block_dict, dict):
        return code_map, label_map
    for k, v in block_dict.items():
        if isinstance(v, dict):
            code = v.get("код", v.get("code", v.get("value", "")))
            label = v.get("кодстрокдоп", v.get("label", ""))
        else:
            code = str(v) if v is not None else ""
            label = ""
        code_map[str(k)] = "" if code is None else str(code)
        label_map[str(k)] = "" if label is None else str(label)
    return code_map, label_map

def parse_range_rules(raw_block):
    rules = []
    def _add(start, end, text, code):
        try:
            start_f = float(start); end_f = float(end)
        except Exception:
            return
        rules.append({"start": start_f, "end": end_f,
                      "text": "" if text is None else str(text),
                      "code": "" if code is None else str(code)})
    if isinstance(raw_block, dict):
        for _, obj in raw_block.items():
            if not isinstance(obj, dict): continue
            start = obj.get("начало", obj.get("start"))
            end   = obj.get("конец",  obj.get("end"))
            text  = obj.get("значение", obj.get("value", obj.get("текст", obj.get("label", ""))))
            code  = obj.get("кодстроки", obj.get("code", obj.get("код")))
            _add(start, end, text, code)
    elif isinstance(raw_block, list):
        for obj in raw_block:
            if not isinstance(obj, dict): continue
            start = obj.get("начало", obj.get("start"))
            end   = obj.get("конец",  obj.get("end"))
            text  = obj.get("значение", obj.get("value", obj.get("текст", obj.get("label", obj.get("ключ", "")))))
            code  = obj.get("кодстроки", obj.get("code", obj.get("код")))
            _add(start, end, text, code)
    rules.sort(key=lambda r: (r["start"], r["end"]))
    return rules

# ====== запись листа (общая функция) ======
def write_df_to_worksheet(ws, final_df, idx_potochna, idx_enddate, text_cols_set,
                          report_every_sec=40, show_progress=False):
    total_rows = final_df.shape[0]
    rows_written = 0
    start_t = time.time()
    last_report = start_t

    for r_idx, row in enumerate(final_df.itertuples(index=False, name=None), start=1):
        for c_idx, val in enumerate(row, start=1):
            # Даты (только на основном листе)
            if (idx_potochna is not None and c_idx == idx_potochna and r_idx >= 2) or \
               (idx_enddate  is not None and c_idx == idx_enddate  and r_idx >= 2):
                if pd.isna(val):
                    py_val = None
                else:
                    py_val = pd.to_datetime(val).to_pydatetime()
                cell = ws.cell(row=r_idx, column=c_idx, value=py_val)
                cell.number_format = "DD.MM.YYYY"
            else:
                # Текст/числа
                if c_idx in text_cols_set:
                    cell = ws.cell(row=r_idx, column=c_idx, value="" if val is None else str(val))
                    cell.number_format = "@"
                else:
                    if isinstance(val, pd.Timestamp):
                        cell = ws.cell(row=r_idx, column=c_idx, value=val.to_pydatetime())
                    elif val == "" or val is None or (isinstance(val, float) and np.isnan(val)):
                        cell = ws.cell(row=r_idx, column=c_idx, value=None)
                    else:
                        cell = ws.cell(row=r_idx, column=c_idx, value=val)

        rows_written += 1
        if show_progress:
            now = time.time()
            if now - last_report >= report_every_sec:
                print(f"Записано строк: {rows_written} из {total_rows}", flush=True)
                last_report = now

    return rows_written, total_rows

# === Основной код ===
if __name__ == "__main__":
    # --- Загружаем app_settings.json (файлы и поточнадата) ---
    if not os.path.exists(SETTINGS_JSON):
        raise FileNotFoundError(f"Не найден файл настроек {SETTINGS_JSON}")
    try:
        with open(SETTINGS_JSON, "r", encoding="utf-8") as f:
            settings = json.load(f)
    except json.JSONDecodeError as e:
        raise SystemExit(f"Ошибка JSON в {SETTINGS_JSON}: {e}")

    # файлы из настроек
    files_from_settings = []
    if isinstance(settings.get("files"), list):
        files_from_settings = [str(p) for p in settings["files"]]
    else:
        # поддержка старого стиля: file1/file2
        for k in ("file1", "file2"):
            if settings.get(k):
                files_from_settings.append(str(settings[k]))

    if not files_from_settings:
        raise ValueError(f"В {SETTINGS_JSON} нужно указать 'files': [\"file1.xlsx\", \"file2.xlsx\"] или 'file1'/'file2'.")

    # sheet_name (опционально)
    SHEET_NAME = settings.get("sheet_name", SHEET_NAME)

    # поточнадата
    date_str = settings.get("поточнадата")
    if not isinstance(date_str, str) or not date_str.strip():
        raise ValueError(f"В {SETTINGS_JSON} нужен ключ 'поточнадата' в формате 'дд.мм.гггг'.")
    try:
        CURRENT_DATE = datetime.strptime(date_str.strip(), "%d.%m.%Y")
    except Exception:
        raise ValueError(f"Значение 'поточнадата' в {SETTINGS_JSON} должно быть в формате 'дд.мм.гггг'.")

    # --- Загружаем status2_map.json (правила/блоки) ---
    if not os.path.exists(CONFIG_JSON):
        raise FileNotFoundError(f"Не найден файл настроек блоков {CONFIG_JSON}")
    with open(CONFIG_JSON, "r", encoding="utf-8") as f:
        config = json.load(f)

    # Блоки настроек
    status2_cfg = config.get("статус2БЛОК", {}) or {}
    status2_cache = {str(k): parse_status2_entry(v) for k, v in status2_cfg.items()}
    s070_cfg = config.get("S070БЛОК", {}) or {}
    s070_code_map, s070_label_map = parse_s070_block(s070_cfg)
    s186_rules = parse_range_rules(config.get("S186БЛОК", []))
    s190_rules = parse_range_rules(config.get("S190БЛОК", []))
    s242_rules = parse_range_rules(config.get("S242БЛОК", []))

    # Какие файлы реально доступны (если найден 1 — работаем с ним)
    available_files = [p for p in files_from_settings if os.path.exists(p)]
    if not available_files:
        raise FileNotFoundError("Не найден ни один из входных файлов из app_settings.json.")

    # Имя результата
    dates = [extract_date_str(os.path.basename(p)) for p in available_files]
    date_part = "_".join(dates) if dates else "unknown"
    OUTPUT_FILE = f"result_{date_part}.xlsx"

    # Заголовки — из первого доступного файла + доп. колонки
    header_source_path = available_files[0]
    headers = get_headers(header_source_path, SHEET_NAME)
    headers += [
        "статус2",
        "поточнадата",
        "датазакинчення",
        "S070Код",
        "S070Строка",
        "S186Строка",
        "S186Код",
        "S186КодиСтрока",
        "S190Строка",
        "S190Код",
        "S190КодИСтрока",
        "СтрокДоПогашення",
        "S242Строка",
        "S242Код",
        "S242КодИСтрока",
        "КомКредСумаУзвітномуперіоді",
    ]

    # Чтение найденных файлов и объединение
    df_list = [read_and_process_one(p, SHEET_NAME) for p in available_files]
    combined = pd.concat(df_list, ignore_index=True) if len(df_list) > 1 else df_list[0].copy()

    # ---- статус2 (с «Прострочений»)
    def compute_status2(row):
        key = row["col_37"]
        if is_empty_like(key):
            try:
                v35 = float(row["col_35"])
            except Exception:
                v35 = 0.0
            if v35 == 0:
                return "Закритий"
            try:
                v38 = float(row["col_38"])
            except Exception:
                v38 = 0.0
            return "Активний" if v38 == 0 else "Прострочений"
        key = str(key)
        if key not in status2_cache:
            return "Ненашли"
        value, need_check38 = status2_cache[key]
        if need_check38:
            try:
                v38 = float(row["col_38"])
            except Exception:
                v38 = 0.0
            return "Активний" if v38 == 0 else "Прострочений"
        return str(value) if value is not None else ""

    combined["статус2"] = combined.apply(compute_status2, axis=1)

    # ---- поточнадата
    combined["поточнадата"] = CURRENT_DATE  # pandas datetime -> Excel date

    # ---- датазакинчення = col_8 + col_9 (дни)
    def compute_end_date(row):
        date_in = row["col_8"]; days = row["col_9"]
        if isinstance(date_in, str) and date_in.strip() != "" and not is_empty_like(days):
            dt = pd.to_datetime(date_in, dayfirst=True, errors="coerce")
            try:
                days_int = int(days)
            except Exception:
                return pd.NaT
            if pd.isna(dt):
                return pd.NaT
            return dt + pd.to_timedelta(days_int, unit="D")
        return pd.NaT
    combined["датазакинчення"] = combined.apply(compute_end_date, axis=1)

    # ---- S070 (код + строка с кодстрокдоп при наличии)
    def map_s070_code(status_val):
        key = "" if is_empty_like(status_val) else str(status_val)
        return s070_code_map.get(key, "00")

    def map_s070_label(status_val):
        key = "" if is_empty_like(status_val) else str(status_val)
        label = s070_label_map.get(key, "")
        return label if label != "" else key

    combined["S070Код"] = combined["статус2"].map(map_s070_code)
    combined["S070Строка"] = combined.apply(
        lambda r: f"{str(r['S070Код']).strip()}-{map_s070_label(r['статус2']).strip()}",
        axis=1
    )

    # ---- S186 (по col_9)
    def compute_s186_pair(row):
        v = row["col_9"]
        try:
            num = float(v)
        except Exception:
            return pd.Series(["00", ""], index=["S186Строка", "S186Код"])
        for rule in s186_rules:
            if rule["start"] <= num <= rule["end"]:
                text = rule.get("text", "").strip()
                code = rule.get("code", "").strip()
                return pd.Series([text if text else "00", code], index=["S186Строка", "S186Код"])
        return pd.Series(["00", ""], index=["S186Строка", "S186Код"])
    combined[["S186Строка", "S186Код"]] = combined.apply(compute_s186_pair, axis=1)
    combined["S186КодиСтрока"] = combined.apply(
        lambda r: f"{str(r['S186Код']).strip()}-{str(r['S186Строка']).strip()}",
        axis=1
    )

    # ---- S190 (по col_38)
    def compute_s190_text(row):
        v = row["col_38"]
        try:
            num = float(v)
        except Exception:
            return "00"
        for rule in s190_rules:
            if rule["start"] <= num <= rule["end"]:
                text = rule.get("text", "").strip()
                return text if text else "00"
        return "00"
    def compute_s190_code(row):
        v = row["col_38"]
        try:
            num = float(v)
        except Exception:
            return "00"
        for rule in s190_rules:
            if rule["start"] <= num <= rule["end"]:
                code = rule.get("code", "").strip()
                return code if code else "00"
        return "00"
    combined["S190Строка"] = combined.apply(compute_s190_text, axis=1)
    combined["S190Код"]    = combined.apply(compute_s190_code,  axis=1)
    combined["S190КодИСтрока"] = combined.apply(
        lambda r: f"{str(r['S190Код']).strip()}-{str(r['S190Строка']).strip()}",
        axis=1
    )

    # ---- СтрокДоПогашення = (датазакинчення - поточнадата) в днях
    combined["СтрокДоПогашення"] = combined.apply(
        lambda r: ("" if pd.isna(r["датазакинчення"]) or pd.isna(r["поточnadата" if False else "поточнадата"])
                   else int((pd.to_datetime(r["датазакинчення"]) - pd.to_datetime(r["поточнадата"])).days)),
        axis=1
    )

    # ---- S242 (по 'СтрокДоПогашення')
    def compute_s242_text(row):
        v = row["СтрокДоПогашення"]
        try:
            num = float(v)
        except Exception:
            return "00"
        for rule in s242_rules:
            if rule["start"] <= num <= rule["end"]:
                text = rule.get("text", "").strip()
                return text if text else "00"
        return "00"
    def compute_s242_code(row):
        v = row["СтрокДоПогашення"]
        try:
            num = float(v)
        except Exception:
            return "00"
        for rule in s242_rules:
            if rule["start"] <= num <= rule["end"]:
                code = rule.get("code", "").strip()
                return code if code else "00"
        return "00"
    combined["S242Строка"] = combined.apply(compute_s242_text, axis=1)
    combined["S242Код"]    = combined.apply(compute_s242_code,  axis=1)
    combined["S242КодИСтрока"] = combined.apply(
        lambda r: f"{str(r['S242Код']).strip()}-{str(r['S242Строка']).strip()}",
        axis=1
    )

    # ---- Новая колонка (Лист1): КомКредСумаУзвітномуперіоді = col_29 - col_13 + col_22 - col_28
    combined["КомКредСумаУзвітномуперіоді"] = (
        combined["col_29"].map(safe_num)
        - combined["col_13"].map(safe_num)
        + combined["col_22"].map(safe_num)
        - combined["col_28"].map(safe_num)
    )

    # ---- CSV вспомогательные
    unique_statuses = sorted(set(v for v in combined["col_37"] if str(v).strip() != ""))
    pd.DataFrame({"статус37": unique_statuses}).to_csv("статусыКолонки37.csv", index=False, encoding="utf-8-sig")

    err_df = combined.loc[combined["статус2"] == "Ненашли", ["col_37", "col_38"]].copy()
    if not err_df.empty:
        err_df.rename(columns={"col_37": "ключ", "col_38": "значение"}, inplace=True)
        err_df = err_df.drop_duplicates(subset=["ключ", "значение"])
        err_df.to_csv("Ошибкистатусов.csv", index=False, encoding="utf-8-sig")
    else:
        pd.DataFrame(columns=["ключ", "значение"]).to_csv("Ошибкистатусов.csv", index=False, encoding="utf-8-sig")

    # ---- Итоговый DataFrame для основного листа (первая строка — заголовки как данные)
    final_df = pd.concat([pd.DataFrame([headers], columns=list(combined.columns)), combined],
                         ignore_index=True)

    # ==== Создаём книгу и листы
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Лист1"
    ws_sel = wb.create_sheet(title="Выборка")
    ws_nbu = wb.create_sheet(title="ДляНБУ")
    ws_fee = wb.create_sheet(title="КомисссияПоКредитамВсе")
    ws_fee_nbu = wb.create_sheet(title="КомисссияПоКредитамНБУ")

    # Индексы дат для форматирования основного листа
    col_idx_map = {name: i+1 for i, name in enumerate(combined.columns)}
    idx_potochna = col_idx_map.get("поточнадата")
    idx_enddate  = col_idx_map.get("датазакинчення")

    # Набор текстовых колонок для основного листа (1-based)
    text_cols_extra = [
        "статус2", "S070Код", "S070Строка",
        "S186Строка", "S186Код", "S186КодиСтрока",
        "S190Строка", "S190Код", "S190КодИСтрока",
        "S242Строка", "S242Код", "S242КодИСтрока",
        # КомКредСумаУзвітномуперіоді — ЧИСЛО
    ]
    text_cols_set = set(TEXT_COLS)
    for nm in text_cols_extra:
        if nm in col_idx_map:
            text_cols_set.add(col_idx_map[nm])

    # Пишем основной лист с прогрессом
    rows_written_1, total_rows_1 = write_df_to_worksheet(
        ws_main, final_df, idx_potochna, idx_enddate, text_cols_set,
        report_every_sec=40, show_progress=True
    )

    # =========================
    #   Листы «Выборка» / «ДляНБУ»
    #   ВАЖНО: берём только строки, где col_14 > 0
    # =========================
    group_keys = [
        "S070Код", "S070Строка",
        "S186Строка", "S186Код", "S186КодиСтрока",
        "S190Строка", "S190Код", "S190КодИСтрока",
        "S242Строка", "S242Код", "S242КодИСтрока",
    ]

    filtered = combined[combined["col_14"].map(safe_num) > 0].copy()
    body_sel = filtered[group_keys + ["col_14"]].copy()
    body_sel.rename(columns={"col_14": "сумма"}, inplace=True)

    agg_full = (body_sel
                .groupby(group_keys, dropna=False, as_index=False)
                .agg(сумма=("сумма", "sum"),
                     колличество=("сумма", "size"))
                )

    # ----- Лист «Выборка» — без доп. фильтра
    headers_sel = group_keys + ["сумма", "колличество"]
    selection_out = pd.concat([pd.DataFrame([headers_sel], columns=headers_sel), agg_full],
                              ignore_index=True)
    sum_idx_sel = selection_out.columns.get_loc("сумма") + 1
    text_cols_sel = set(range(1, len(selection_out.columns) + 1)) - {sum_idx_sel}
    _rw_sel, _tr_sel = write_df_to_worksheet(
        ws_sel, selection_out, idx_potochna=None, idx_enddate=None, text_cols_set=text_cols_sel,
        report_every_sec=40, show_progress=False
    )

    # ----- Лист «ДляНБУ» — плюс фильтр сумма>0 и колличество>0
    agg_nbu = agg_full[(agg_full["сумма"] > 0) & (agg_full["колличество"] > 0)].copy()
    selection_nbu = pd.concat([pd.DataFrame([headers_sel], columns=headers_sel), agg_nbu],
                              ignore_index=True)
    sum_idx_nbu = selection_nbu.columns.get_loc("сумма") + 1
    text_cols_nbu = set(range(1, len(selection_nbu.columns) + 1)) - {sum_idx_nbu}
    _rw_nbu, _tr_nbu = write_df_to_worksheet(
        ws_nbu, selection_nbu, idx_potochna=None, idx_enddate=None, text_cols_set=text_cols_nbu,
        report_every_sec=40, show_progress=False
    )

    # ----- Лист «КомисссияПоКредитамВсе»
    fee_keys = [
        "S070Код", "S070Строка",
        "S190Строка", "S190Код", "S190КодИСтрока",
    ]
    fee_base = combined[fee_keys + ["КомКредСумаУзвітномуперіоді", "col_34"]].copy()
    fee_base.rename(columns={"col_34": "СумаНазвітнудату"}, inplace=True)

    fee_agg = (fee_base
               .groupby(fee_keys, dropna=False, as_index=False)
               .agg(КомКредСумаУзвітномуперіоді=("КомКредСумаУзвітномуперіоді", "sum"),
                    СумаНазвітнудату=("СумаНазвітнудату", "sum"),
                    колличество=("КомКредСумаУзвітномуперіоді", "size"))
               )

    headers_fee = fee_keys + ["КомКредСумаУзвітномуперіоді", "СумаНазвітнудату", "колличество"]
    fee_out = pd.concat([pd.DataFrame([headers_fee], columns=headers_fee), fee_agg],
                        ignore_index=True)

    # На этом листе ЧИСЛОВЫЕ: КомКредСумаУзвітномуперіоді и СумаНазвітнудату
    num_idx_fee_1 = fee_out.columns.get_loc("КомКредСумаУзвітномуперіоді") + 1
    num_idx_fee_2 = fee_out.columns.get_loc("СумаНазвітнудату") + 1
    text_cols_fee = set(range(1, len(fee_out.columns) + 1)) - {num_idx_fee_1, num_idx_fee_2}
    _rw_fee, _tr_fee = write_df_to_worksheet(
        ws_fee, fee_out, idx_potochna=None, idx_enddate=None, text_cols_set=text_cols_fee,
        report_every_sec=40, show_progress=False
    )

    # ----- Лист «КомисссияПоКредитамНБУ» (без «колличество», с фильтром нулевых сумм)
    fee_nbu = (fee_agg[fee_keys + ["КомКредСумаУзвітномуперіоді", "СумаНазвітнудату"]]
               .groupby(fee_keys, dropna=False, as_index=False)
               .agg(КомКредСумаУзвітномуперіоді=("КомКредСумаУзвітномуперіоді", "sum"),
                    СумаНазвітнудату=("СумаНазвітнудату", "sum"))
               )
    fee_nbu = fee_nbu[
        (fee_nbu["КомКредСумаУзвітномуперіоді"] != 0) |
        (fee_nbu["СумаНазвітнудату"] != 0)
    ]

    headers_fee_nbu = fee_keys + ["КомКредСумаУзвітномуперіоді", "СумаНазвітнудату"]
    fee_nbu_out = pd.concat([pd.DataFrame([headers_fee_nbu], columns=headers_fee_nbu), fee_nbu],
                            ignore_index=True)

    num_idx_fee_nbu_1 = fee_nbu_out.columns.get_loc("КомКредСумаУзвітномуперіоді") + 1
    num_idx_fee_nbu_2 = fee_nbu_out.columns.get_loc("СумаНазвітнудату") + 1
    text_cols_fee_nbu = set(range(1, len(fee_nbu_out.columns) + 1)) - {num_idx_fee_nbu_1, num_idx_fee_nbu_2}
    _rw_fee_nbu, _tr_fee_nbu = write_df_to_worksheet(
        ws_fee_nbu, fee_nbu_out, idx_potochna=None, idx_enddate=None, text_cols_set=text_cols_fee_nbu,
        report_every_sec=40, show_progress=False
    )

    # Сохраняем книгу
    wb.save(OUTPUT_FILE)

    # Финал
    print(f"Готово. Записано строк (лист1): {rows_written_1} из {total_rows_1}. Колонок (лист1): {final_df.shape[1]}")
    print(f"Excel сохранён: {os.path.abspath(OUTPUT_FILE)}")
    print("Созданы CSV: статусыКолонки37.csv и Ошибкистатусов.csv")
    print("Обработаны файлы:", ", ".join(available_files))
    print(f"Лист 'Выборка': строк {selection_out.shape[0]}, колонок {selection_out.shape[1]}")
    print(f"Лист 'ДляНБУ': строк {selection_nbu.shape[0]}, колонок {selection_nbu.shape[1]}")
    print(f"Лист 'КомисссияПоКредитамВсе': строк {fee_out.shape[0]}, колонок {fee_out.shape[1]}")
    print(f"Лист 'КомисссияПоКредитамНБУ': строк {fee_nbu_out.shape[0]}, колонок {fee_nbu_out.shape[1]}")
